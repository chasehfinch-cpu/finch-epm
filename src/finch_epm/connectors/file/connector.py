"""File connector for CSV and Excel import.

Imports local files directly into the DuckDB cache. Does not implement
the full ConnectorBase interface (no connect/introspect) -- instead
provides direct import functions usable from the CLI and the sync engine.

This is not a database connector. It reads files and loads them as
tables in the local cache, making them queryable in .fdash dashboards
alongside NetSuite and SQL Server data.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from finch_epm.cache.local import LocalCacheEngine
from finch_epm.connectors.types import ColumnType

logger = logging.getLogger(__name__)


def import_csv(
    file_path: str | Path,
    cache: LocalCacheEngine,
    table_name: str | None = None,
    delimiter: str = ",",
    has_header: bool = True,
) -> int:
    """Import a CSV file into the local DuckDB cache as a table.

    Args:
        file_path: Path to the CSV file.
        cache: The cache engine to import into.
        table_name: Table name in the cache. Defaults to the filename
            without extension.
        delimiter: CSV delimiter (default comma).
        has_header: Whether the first row is a header.

    Returns:
        Number of rows imported.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if table_name is None:
        table_name = path.stem.replace(" ", "_").replace("-", "_")

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)

        if has_header:
            header = next(reader)
            column_names = [_sanitize_column(h) for h in header]
        else:
            first_row = next(reader)
            column_names = [f"col_{i}" for i in range(len(first_row))]
            rows = [first_row]
            for row in reader:
                rows.append(row)
            column_types = ["string"] * len(column_names)
            return cache.ingest_facts(
                table_name, column_names, column_types, rows, mode="replace"
            )

        rows: list[list[Any]] = []
        for row in reader:
            # Pad or truncate to match header length
            padded = row[:len(column_names)]
            while len(padded) < len(column_names):
                padded.append(None)
            rows.append(padded)

    column_types = ["string"] * len(column_names)
    return cache.ingest_facts(
        table_name, column_names, column_types, rows, mode="replace"
    )


def import_excel(
    file_path: str | Path,
    cache: LocalCacheEngine,
    sheet_name: str | None = None,
    table_name: str | None = None,
) -> int:
    """Import an Excel worksheet into the local DuckDB cache as a table.

    Requires openpyxl to be installed.

    Args:
        file_path: Path to the .xlsx file.
        cache: The cache engine to import into.
        sheet_name: Worksheet name. Defaults to the active sheet.
        table_name: Table name in the cache. Defaults to the sheet name
            or filename.

    Returns:
        Number of rows imported.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel import. "
            "Install with: pip install openpyxl"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if table_name is None:
        table_name = _sanitize_column(ws.title or path.stem)

    # Read all rows
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return 0

    # First row is header
    header = all_rows[0]
    column_names = [_sanitize_column(str(h)) if h else f"col_{i}" for i, h in enumerate(header)]

    rows: list[list[Any]] = []
    for row in all_rows[1:]:
        # Convert all values to strings (DuckDB will handle them)
        converted = []
        for val in row[:len(column_names)]:
            if val is None:
                converted.append(None)
            else:
                converted.append(str(val))
        while len(converted) < len(column_names):
            converted.append(None)
        rows.append(converted)

    column_types = ["string"] * len(column_names)
    return cache.ingest_facts(
        table_name, column_names, column_types, rows, mode="replace"
    )


def list_excel_sheets(file_path: str | Path) -> list[str]:
    """List all worksheet names in an Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def _sanitize_column(name: str) -> str:
    """Sanitize a column name for DuckDB compatibility."""
    # Replace spaces and special chars with underscores
    sanitized = name.strip()
    for char in " -/\\()[]{}.,;:!@#$%^&*+=|<>?'\"":
        sanitized = sanitized.replace(char, "_")
    # Remove consecutive underscores
    while "__" in sanitized:
        sanitized = sanitized.replace("__", "_")
    # Remove leading/trailing underscores
    sanitized = sanitized.strip("_")
    return sanitized.lower() if sanitized else "unnamed"
